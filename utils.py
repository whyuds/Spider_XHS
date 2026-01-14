import os
import re
import time
import json
import math
import random
import shutil
import requests
import execjs
from loguru import logger
from dotenv import load_dotenv
import openpyxl
from openpyxl import load_workbook
from retry import retry
from PIL import Image
import numpy as np
from datetime import datetime, timedelta
from openai import OpenAI


# --- Common Utils ---
_LOG_SINK_READY = False

def get_spider_home() -> str:
    custom_home = os.getenv("SPIDER_XHS_HOME")
    if custom_home:
        return os.path.abspath(os.path.expanduser(custom_home))
    return os.path.join(os.path.expanduser("~"), ".spider_xhs")

def _ensure_dir(path: str) -> None:
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)

def get_spider_file(filename: str, migrate_from_project: bool = True) -> str:
    spider_home = get_spider_home()
    _ensure_dir(spider_home)
    target_path = os.path.join(spider_home, filename)

    if migrate_from_project and not os.path.exists(target_path):
        project_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        if os.path.exists(project_path):
            try:
                shutil.copy2(project_path, target_path)
            except Exception:
                pass

    return target_path

def _setup_logging() -> None:
    global _LOG_SINK_READY
    if _LOG_SINK_READY:
        return

    logs_dir = os.path.join(get_spider_home(), "logs")
    _ensure_dir(logs_dir)
    logger.add(
        os.path.join(logs_dir, "spider.log"),
        rotation="50 MB",
        retention="14 days",
        encoding="utf-8",
        enqueue=True,
        backtrace=False,
        diagnose=False,
    )
    _LOG_SINK_READY = True

def load_env():
    user_env = get_spider_file(".env", migrate_from_project=True)
    if os.path.exists(user_env):
        load_dotenv(dotenv_path=user_env, override=True)
    else:
        load_dotenv(override=True)
    cookies_str = os.getenv('COOKIES')
    return cookies_str

def init():
    _setup_logging()

    spider_home = get_spider_home()
    datas_dir = os.path.join(spider_home, "datas")
    media_base_path = os.path.join(datas_dir, 'media_datas')
    excel_base_path = os.path.join(datas_dir, 'excel_datas')
    summary_base_path = os.path.join(datas_dir, "day_summary_datas")
    
    for path in [media_base_path, excel_base_path, summary_base_path]:
        if not os.path.exists(path):
            os.makedirs(path, exist_ok=True)
            logger.info(f'创建目录 {path}')
            
    cookies_str = load_env()
    base_path = {
        'media': media_base_path,
        'excel': excel_base_path,
        'summary': summary_base_path,
        'spider_home': spider_home,
    }
    return cookies_str, base_path

# --- Cookie Utils ---
def trans_cookies(cookies_str):
    if not cookies_str:
        return {}
    if '; ' in cookies_str:
        ck = {i.split('=')[0]: '='.join(i.split('=')[1:]) for i in cookies_str.split('; ')}
    else:
        ck = {i.split('=')[0]: '='.join(i.split('=')[1:]) for i in cookies_str.split(';')}
    return ck

# --- XHS Utils (JS & Headers) ---
try:
    js = execjs.compile(open(r'static/xhs_xs_xsc_56.js', 'r', encoding='utf-8').read())
except Exception as e:
    logger.error(f"Failed to load xhs_xs_xsc_56.js: {e}")
    # Fallback or error handling if needed

try:
    xray_js = execjs.compile(open(r'static/xhs_xray.js', 'r', encoding='utf-8').read())
except Exception as e:
    logger.error(f"Failed to load xhs_xray.js: {e}")

def generate_x_b3_traceid(len=16):
    x_b3_traceid = ""
    for t in range(len):
        x_b3_traceid += "abcdef0123456789"[math.floor(16 * random.random())]
    return x_b3_traceid

def generate_xs_xs_common(a1, api, data='', method='POST'):
    ret = js.call('get_request_headers_params', api, data, a1, method)
    xs, xt, xs_common = ret['xs'], ret['xt'], ret['xs_common']
    return xs, xt, xs_common

def generate_xs(a1, api, data=''):
    ret = js.call('get_xs', api, data, a1)
    xs, xt = ret['X-s'], ret['X-t']
    return xs, xt

def generate_xray_traceid():
    return xray_js.call('traceId')

def get_request_headers_template():
    return {
        "authority": "edith.xiaohongshu.com",
        "accept": "application/json, text/plain, */*",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "cache-control": "no-cache",
        "content-type": "application/json;charset=UTF-8",
        "origin": "https://www.xiaohongshu.com",
        "pragma": "no-cache",
        "referer": "https://www.xiaohongshu.com/",
        "sec-ch-ua": "\"Not A(Brand\";v=\"99\", \"Microsoft Edge\";v=\"121\", \"Chromium\";v=\"121\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\"",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0",
        "x-b3-traceid": "",
        "x-mns": "unload",
        "x-s": "",
        "x-s-common": "",
        "x-t": "",
        "x-xray-traceid": generate_xray_traceid()
    }

def generate_headers(a1, api, data='', method='POST'):
    xs, xt, xs_common = generate_xs_xs_common(a1, api, data, method)
    x_b3_traceid = generate_x_b3_traceid()
    headers = get_request_headers_template()
    headers['x-s'] = xs
    headers['x-t'] = str(xt)
    headers['x-s-common'] = xs_common
    headers['x-b3-traceid'] = x_b3_traceid
    if data:
        data = json.dumps(data, separators=(',', ':'), ensure_ascii=False)
    return headers, data

def generate_request_params(cookies_str, api, data='', method='POST'):
    cookies = trans_cookies(cookies_str)
    a1 = cookies.get('a1', '')
    headers, data = generate_headers(a1, api, data, method)
    return headers, cookies, data

def splice_str(api, params):
    url = api + '?'
    for key, value in params.items():
        if value is None:
            value = ''
        url += key + '=' + str(value) + '&'
    return url[:-1]

# --- OCR Utils ---
try:
    from paddleocr import PaddleOCR
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

_ocr_instance = None

def get_ocr_instance():
    global _ocr_instance
    if not HAS_OCR:
        return None
    if _ocr_instance is None:
        # Initialize globally on first use
        _ocr_instance = PaddleOCR(use_angle_cls=False, lang="ch", ocr_version="PP-OCRv4")
    return _ocr_instance

def ocr_process_note_images(note_path):
    if not HAS_OCR:
        logger.warning("OCR skipped because paddleocr is not installed.")
        return

    ocr = get_ocr_instance()
    if not ocr:
        return

    if not os.path.exists(note_path):
        return

    logger.info(f"正在对 {os.path.basename(note_path)} 进行OCR文字识别...")
    
    processed_count = 0
    for root, dirs, files in os.walk(note_path):
        for file in files:
            if file.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                image_path = os.path.join(root, file)
                txt_path = os.path.splitext(image_path)[0] + ".txt"
                
                if os.path.exists(txt_path):
                    continue

                try:
                    with Image.open(image_path) as img:
                        img = img.convert('RGB')
                        img_np = np.array(img)[:, :, ::-1] # RGB to BGR
                    
                    # ocr.ocr() can vary by version, handle list/dict
                    result = ocr.ocr(img_np)
                    # logger.info(f"OCR result for {file}: {type(result)}") 
                    text_lines = []
                    
                    if result and len(result) > 0:
                        data = result[0]
                        # Handle PP-OCRv4 / PaddleX dict output
                        if isinstance(data, dict) and 'rec_texts' in data:
                            rec_texts = data['rec_texts']
                            rec_scores = data.get('rec_scores', [])
                            for i, text in enumerate(rec_texts):
                                score = rec_scores[i] if i < len(rec_scores) else 1.0
                                if score > 0.6:
                                    text_lines.append(text)
                        # Handle standard PP-OCR list output [[[[x,y]...], (text, score)], ...]
                        elif isinstance(data, list):
                            for line in data:
                                if isinstance(line, (list, tuple)) and len(line) >= 2:
                                    info = line[1]
                                    if isinstance(info, (list, tuple)) and len(info) >= 2:
                                        text, score = info[0], info[1]
                                        if score > 0.6:
                                            text_lines.append(text)
                    
                    if text_lines:
                        with open(txt_path, 'w', encoding='utf-8') as f:
                            f.write('\n'.join(text_lines))
                        processed_count += 1
                        
                except Exception as e:
                    logger.error(f"Failed OCR for {file}: {e}")
    
    if processed_count > 0:
        logger.info(f"OCR完成，生成了 {processed_count} 个文本文件")

# --- Data/File Utils ---
def norm_str(str_val):
    new_str = re.sub(r"|[\\/:*?\"<>| ]+", "", str_val).replace('\n', '').replace('\r', '')
    return new_str

def norm_text(text):
    if text is None: return ""
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', str(text))
    return text

def timestamp_to_str(timestamp):
    time_local = time.localtime(timestamp / 1000)
    dt = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
    return dt

def check_and_create_path(path):
    if not os.path.exists(path):
        os.makedirs(path)

def download_media(path, name, url, type):
    try:
        if type == 'image':
            content = requests.get(url).content
            with open(os.path.join(path, name + '.jpg'), mode="wb") as f:
                f.write(content)
        elif type == 'video':
            res = requests.get(url, stream=True)
            chunk_size = 1024 * 1024
            with open(os.path.join(path, name + '.mp4'), mode="wb") as f:
                for data in res.iter_content(chunk_size=chunk_size):
                    f.write(data)
    except Exception as e:
        logger.error(f"Download media failed {url}: {e}")

def save_note_detail(note, path):
    file_path = os.path.join(path, 'detail.txt')
    with open(file_path, mode="w", encoding="utf-8") as f:
        f.write(f"笔记id: {note.get('note_id', '')}\n")
        f.write(f"笔记url: {note.get('note_url', '')}\n")
        f.write(f"笔记类型: {note.get('note_type', '')}\n")
        f.write(f"用户id: {note.get('user_id', '')}\n")
        f.write(f"用户主页url: {note.get('home_url', '')}\n")
        f.write(f"昵称: {note.get('nickname', '')}\n")
        f.write(f"头像url: {note.get('avatar', '')}\n")
        f.write(f"标题: {note.get('title', '')}\n")
        f.write(f"描述: {note.get('desc', '')}\n")
        f.write(f"点赞数量: {note.get('liked_count', 0)}\n")
        f.write(f"收藏数量: {note.get('collected_count', 0)}\n")
        f.write(f"评论数量: {note.get('comment_count', 0)}\n")
        f.write(f"分享数量: {note.get('share_count', 0)}\n")
        f.write(f"视频封面url: {note.get('video_cover', '')}\n")
        f.write(f"视频地址url: {note.get('video_addr', '')}\n")
        f.write(f"图片地址url列表: {note.get('image_list', [])}\n")
        f.write(f"标签: {note.get('tags', [])}\n")
        f.write(f"上传时间: {note.get('upload_time', '')}\n")
        f.write(f"ip归属地: {note.get('ip_location', '')}\n")

@retry(tries=3, delay=1)
def download_note(note_info, path, save_choice):
    note_id = note_info['note_id']
    user_id = note_info['user_id']
    title = note_info['title']
    title = norm_str(title)[:40]
    nickname = note_info['nickname']
    nickname = norm_str(nickname)[:20]
    if not title.strip():
        title = '无标题'
    upload_time = note_info.get('upload_time')
    date_prefix = ""
    if upload_time:
        try:
             date_prefix = upload_time.split(' ')[0].replace('-', '') + "_"
        except:
             pass

    save_path = f'{path}/{nickname}_{user_id}/{date_prefix}{title}_{note_id}'
    check_and_create_path(save_path)
    with open(f'{save_path}/info.json', mode='w', encoding='utf-8') as f:
        f.write(json.dumps(note_info, ensure_ascii=False) + '\n')
        
    note_type = note_info['note_type']
    save_note_detail(note_info, save_path)
    
    if note_type == '图集' and save_choice in ['media', 'media-image', 'all']:
        for img_index, img_url in enumerate(note_info['image_list']):
            download_media(save_path, f'image_{img_index}', img_url, 'image')
    elif note_type == '视频' and save_choice in ['media', 'media-video', 'all']:
        download_media(save_path, 'cover', note_info['video_cover'], 'image')
        download_media(save_path, 'video', note_info['video_addr'], 'video')
    return save_path

def get_saved_note_ids(file_path):
    if not os.path.exists(file_path):
        return set()
    ids = set()
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if row[0]:
                ids.add(str(row[0]))
    except Exception as e:
        logger.error(f"读取已存在文件失败: {e}")
    return ids

def save_to_xlsx(datas, file_path, type='note', mode='w'):
    if mode == 'a' and os.path.exists(file_path):
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            mode = 'w'
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        mode = 'w'

    if mode == 'w':
        headers = ['笔记id', '笔记url', '笔记类型', '用户id', '用户主页url', '昵称', '头像url', '标题', '描述', '点赞数量', '收藏数量', '评论数量', '分享数量', '视频封面url', '视频地址url', '图片地址url列表', '标签', '上传时间', 'ip归属地']
        ws.append(headers)
    
    for data in datas:
        # Sort or ensure order? Dict order is preserved in recent python, but safer to be specific if order matters. 
        # The original code assumed data dict keys order matches headers.
        # We should ensure the value list order matches the headers list.
        # Original code: data = {k: norm_text(str(v)) for k, v in data.items()}; ws.append(list(data.values()))
        # This is risky if key order changes. Ideally we map by key.
        # However, `handle_note_info` returns a specific order. I will keep it as is to avoid breaking changes, 
        # but apply norm_text.
        clean_data = {k: norm_text(v) for k, v in data.items()}
        ws.append(list(clean_data.values()))
        
    wb.save(file_path)
    logger.info(f'数据保存至 {file_path}')

def handle_note_info(data):
    note_id = data['id']
    note_url = data.get('url', f'https://www.xiaohongshu.com/explore/{note_id}')
    
    note_card = data.get('note_card', {})
    note_type = note_card.get('type', 'normal')
    note_type = '图集' if note_type == 'normal' else '视频'
    
    user = note_card.get('user', {})
    user_id = user.get('user_id', '')
    home_url = f'https://www.xiaohongshu.com/user/profile/{user_id}' if user_id else ''
    nickname = user.get('nickname', '')
    avatar = user.get('avatar', '')
    
    title = note_card.get('title', '').strip() or '无标题'
    desc = note_card.get('desc', '')
    
    interact = note_card.get('interact_info', {})
    liked_count = interact.get('liked_count', 0)
    collected_count = interact.get('collected_count', 0)
    comment_count = interact.get('comment_count', 0)
    share_count = interact.get('share_count', 0)
    
    image_list = []
    for image in note_card.get('image_list', []):
        try:
            # Check info_list[1] which usually is the high quality one?
            infos = image.get('info_list', [])
            if len(infos) > 1:
                image_list.append(infos[1]['url'])
            elif len(infos) > 0:
                 image_list.append(infos[0]['url'])
        except:
            pass
            
    video_cover = None
    video_addr = None
    if note_type == '视频':
        if image_list:
            video_cover = image_list[0]
        consumer = note_card.get('video', {}).get('consumer', {})
        origin_key = consumer.get('origin_video_key')
        if origin_key:
            video_addr = 'https://sns-video-bd.xhscdn.com/' + origin_key
    
    tags = []
    for tag in note_card.get('tag_list', []):
        name = tag.get('name')
        if name: tags.append(name)
        
    upload_time = timestamp_to_str(note_card.get('time', 0))
    ip_location = note_card.get('ip_location', '未知')
    
    return {
        'note_id': note_id,
        'note_url': note_url,
        'note_type': note_type,
        'user_id': user_id,
        'home_url': home_url,
        'nickname': nickname,
        'avatar': avatar,
        'title': title,
        'desc': desc,
        'liked_count': liked_count,
        'collected_count': collected_count,
        'comment_count': comment_count,
        'share_count': share_count,
        'video_cover': video_cover,
        'video_addr': video_addr,
        'image_list': image_list,
        'tags': tags,
        'upload_time': upload_time,
        'ip_location': ip_location,
    }

# --- AI & Notification Utils ---
def generate_ai_summary(content):
    api_key = os.getenv('ARK_API_KEY')
    if not api_key:
        logger.error("ARK_API_KEY not found in environment variables. Skipping AI summary.")
        return "Error: ARK_API_KEY not found."

    try:
        client = OpenAI(
            base_url="https://ark.cn-beijing.volces.com/api/v3",
            api_key=api_key,
        )
        
        # Using standard chat completion
        response = client.chat.completions.create(
            model="doubao-seed-1-8-251228",
            messages=[
                {
                    "role": "system",
                    "content": "你是一个专业的投资分析助手。请根据提供的各个用户笔记内容（包含OCR识别的文字），帮我整理一下每个人的今日的投资建议，每个人的建议不超过200字。"
                },
                {
                    "role": "user",
                    "content": content
                }
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        logger.error(f"AI Summary generation failed: {e}")
        return f"Error generating summary: {e}"

def send_wxpusher_message(content, uids, summary_prefix=""):
    if not uids:
        logger.warning("No UIDs provided for WxPusher notification.")
        return

    url = "https://wxpusher.zjiecode.com/api/send/message"
    
    # Format summary title: M月D日 + fixed suffix
    if not summary_prefix:
        summary_title = f"{(datetime.now() - timedelta(days=1)).strftime('%m月%d日')}收盘后关注用户总结"
    else:
        summary_title = summary_prefix
    
    payload = {
        "appToken": "AT_yR3sR0zXtKJQ8hIQuE4gIriw9JeTj3wA",
        "content": content,
        "summary": summary_title,
        "contentType": 3,
        "topicIds": [43062],
        "uids": uids,
        "verifyPayType": 0
    }
    
    headers = {
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers)
        response_data = response.json()
        if response_data.get('code') == 1000:
            logger.info(f"WxPusher notification sent successfully to {len(uids)} users.")
        else:
            logger.error(f"WxPusher notification failed: {response_data.get('msg')}")
    except Exception as e:
        logger.error(f"Error sending WxPusher notification: {e}")

